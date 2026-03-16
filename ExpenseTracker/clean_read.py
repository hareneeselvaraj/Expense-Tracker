import openpyxl, codecs, sqlite3

XLSX_PATH = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\TestData.xlsx'
DB_PATH = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\ExpenseTracker\expense.db'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

with codecs.open('output_read.txt', 'w', 'utf-8') as f:
    f.write(f"Headers:\n")
    headers = [str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    f.write(str(headers) + "\n\nData:\n")
    
    for r in range(2, min(ws.max_row + 1, 8)):
        row_data = [str(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        f.write(str(row_data) + "\n")
        
    f.write("\nUsers:\n")
    conn = sqlite3.connect(DB_PATH)
    for r in conn.execute("SELECT Id, Name, Email FROM Users"):
        f.write(f"{r[0]} | {r[1]} | {r[2]}\n")

print("Done")
