"""
Read TestData.xlsx and import into the ExpenseTracker database.
Uses openpyxl to read .xlsx files.
"""
import sqlite3
import uuid
import re
import os
from datetime import datetime

# Try openpyxl first, fall back to csv
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

XLSX_PATH = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\TestData.xlsx'
DB_PATH = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\ExpenseTracker\expense.db'

# First, let's just read what's in the file
if HAS_OPENPYXL:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    print()
    
    # Print headers (row 1)
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else '')
    print(f"Headers: {headers}")
    print()
    
    # Print first 5 data rows
    for row in range(2, min(ws.max_row + 1, 7)):
        vals = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            vals.append(str(v) if v is not None else '')
        print(f"  Row {row}: {vals}")
else:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    print("Installed. Please re-run this script.")

# Also check which users exist
print("\n=== DB Users ===")
conn = sqlite3.connect(DB_PATH)
for r in conn.execute("SELECT Id, Name, Email FROM Users"):
    print(f"  {r[0]}  |  {r[1]}  |  {r[2]}")
conn.close()
