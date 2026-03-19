"""
Transaction Importer with AI-Powered Categorization
Reads Excel (.xlsx) bank statements & inserts into ExpenseTracker SQLite DB.
"""

import sqlite3
import uuid
import re
import sys
import os
from datetime import datetime
import openpyxl

# ─── CONFIG ───────────────────────────────────────────────────────────
DB_PATH = r'c:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\ExpenseTracker\expense.db'
USER_ID = '33A9C1A2-AF41-4962-AB67-1C4DC7C6E4D1'  # test2 user
DEFAULT_ACCOUNT_ID = 'B30B43E9-29E8-46BC-B518-4E5CC38B550F'  # HDFC debit account

# ─── AI CATEGORIZATION ENGINE ─────────────────────────────────────────
CATEGORY_RULES = [
    {'keywords': ['swiggy', 'zomato', 'food', 'restaurant', 'cafe', 'pizza', 'burger', 'chicken', 'biryani', 'egg', 'milk', 'bread', 'tea', 'coffee', 'hotel', 'mess', 'canteen', 'dine', 'eat', 'liver', 'mutton', 'fish', 'rice'], 'category_name': 'Food', 'type': 'Expense'},
    {'keywords': ['uber', 'ola', 'rapido', 'petrol', 'diesel', 'fuel', 'metro', 'bus', 'train', 'irctc', 'redbus', 'cab', 'auto', 'parking', 'toll', 'fastag'], 'category_name': 'Transportation', 'type': 'Expense'},
    {'keywords': ['amazon', 'flipkart', 'myntra', 'ajio', 'meesho', 'shopping', 'mall', 'store', 'mart', 'retail', 'reliance', 'dmart', 'bigbasket', 'blinkit', 'instamart', 'zepto'], 'category_name': 'Shopping', 'type': 'Expense'},
    {'keywords': ['rent', 'electricity', 'water', 'gas', 'bill', 'maintenance', 'society', 'broadband', 'wifi', 'internet', 'airtel', 'jio', 'vodafone', 'bsnl', 'recharge', 'dth', 'tata'], 'category_name': 'Rent & Bills', 'type': 'Expense'},
    {'keywords': ['netflix', 'hotstar', 'prime', 'spotify', 'youtube', 'movie', 'theatre', 'game', 'play', 'bookmyshow', 'pvr', 'inox'], 'category_name': 'Entertainment', 'type': 'Expense'},
    {'keywords': ['hospital', 'doctor', 'pharma', 'medical', 'medicine', 'apollo', 'medplus', 'netmed', 'health', 'gym', 'fitness', 'clinic'], 'category_name': 'Health', 'type': 'Expense'},
    {'keywords': ['college', 'university', 'school', 'tuition', 'course', 'udemy', 'coursera', 'book', 'exam', 'fees'], 'category_name': 'Education', 'type': 'Expense'},
    {'keywords': ['trade', 'eba', 'eq trade', 'mutual fund', 'sip', 'zerodha', 'groww', 'upstox', 'angel', 'nse', 'bse', 'share', 'dividend', 'stock'], 'category_name': 'Investment', 'type': 'Expense', 'override_type': 'Investment'},
    {'keywords': ['salary', 'credit', 'refund', 'cashback', 'reward', 'interest', 'dividend'], 'category_name': 'Salary', 'type': 'Income'},
]

def generate_smart_description(raw_remarks):
    if not raw_remarks: return "Transaction"
    remarks = raw_remarks.strip()
    upi_match = re.match(r'UPI/([^/]+)/([^/]+)/([^/]*)/([^/]*)', remarks, re.IGNORECASE)
    if upi_match:
        merchant = upi_match.group(1).strip().title()
        purpose = upi_match.group(3).strip().title() if upi_match.group(3) else ''
        return f"UPI Payment to {merchant} - {purpose}" if purpose else f"UPI Payment to {merchant}"
    if remarks.upper().startswith('EBA/'):
        trade_match = re.match(r'EBA/EQ Trade (\d{2}\w{3})/(\d+)', remarks)
        return f"Equity Trade on {trade_match.group(1)}" if trade_match else "Equity/Stock Trade"
    neft_match = re.match(r'(NEFT|RTGS|IMPS)/([^/]+)', remarks, re.IGNORECASE)
    if neft_match: return f"{neft_match.group(1).upper()} Transfer - {neft_match.group(2).strip().title()}"
    if 'ATM' in remarks.upper(): return "ATM Cash Withdrawal"
    if 'salary' in remarks.lower(): return "Salary Credit"
    cleaned = re.sub(r'[A-Fa-f0-9]{20,}', '', remarks)
    cleaned = re.sub(r'/+', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned[:120] if len(cleaned) >= 3 else remarks[:100]

def categorize_transaction(remarks, is_deposit):
    if not remarks: return ('Salary', 'Income') if is_deposit else ('Shopping', 'Expense')
    remarks_lower = remarks.lower()
    if is_deposit: return ('Salary', 'Income')
    for rule in CATEGORY_RULES:
        for kw in rule['keywords']:
            if kw in remarks_lower:
                return (rule['category_name'], rule.get('override_type', rule['type']))
    return ('Shopping', 'Expense')

def get_category_id(conn, user_id, category_name, category_type):
    row = conn.execute("SELECT Id FROM Categories WHERE UserId = ? AND Name = ? AND Type = ? LIMIT 1", (user_id, category_name, category_type)).fetchone()
    if row: return row[0]
    row = conn.execute("SELECT Id FROM Categories WHERE UserId = ? AND Name = ? LIMIT 1", (user_id, category_name)).fetchone()
    if row: return row[0]
    row = conn.execute("SELECT Id FROM Categories WHERE UserId = ? AND Type = ? LIMIT 1", (user_id, category_type)).fetchone()
    if row: return row[0]
    # If NO CATEGORY AT ALL, create one
    cat_id = str(uuid.uuid4()).upper()
    conn.execute("INSERT INTO Categories (Id, Name, Type, Icon, Color, UserId) VALUES (?, ?, ?, ?, ?, ?)", (cat_id, category_name, category_type, 'HelpCircle', '#cccccc', user_id))
    print(f"  [AI] Created missing category: {category_name} ({category_type})")
    return cat_id

def parse_xlsx(file_path):
    transactions = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Map headers
    headers = [str(ws.cell(row=1, column=c).value).lower() for c in range(1, ws.max_column + 1)]
    date_col = next((i for i, h in enumerate(headers) if 'date' in h), None)
    remark_col = next((i for i, h in enumerate(headers) if 'remark' in h), None)
    with_col = next((i for i, h in enumerate(headers) if 'withdrawal' in h), None)
    dep_col = next((i for i, h in enumerate(headers) if 'deposit' in h), None)
    
    for r in range(2, ws.max_row + 1):
        # 1. Date
        date_val = str(ws.cell(row=r, column=date_col + 1).value) if date_col is not None else None
        if not date_val or date_val == 'None' or not date_val.strip(): continue
        
        date_obj = None
        for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                date_obj = datetime.strptime(date_val.strip(), fmt)
                break
            except ValueError: pass
            
        if not date_obj: continue
        
        # 2. Remarks
        remarks = ws.cell(row=r, column=remark_col + 1).value if remark_col is not None else ''
        remarks = str(remarks).strip() if remarks else ''
        
        # 3. Amount
        withdrawal_val = ws.cell(row=r, column=with_col + 1).value if with_col is not None else None
        deposit_val = ws.cell(row=r, column=dep_col + 1).value if dep_col is not None else None
        
        withdrawal = float(withdrawal_val) if withdrawal_val else 0
        deposit = float(deposit_val) if deposit_val else 0
            
        is_deposit = deposit > 0
        amount = deposit if is_deposit else withdrawal
        
        if amount <= 0: continue
        
        cat_name, tx_type = categorize_transaction(remarks, is_deposit)
        description = generate_smart_description(remarks)
        
        transactions.append({
            'date': date_obj.strftime('%Y-%m-%dT00:00:00'),
            'amount': round(amount, 2),
            'type': tx_type,
            'category_name': cat_name,
            'description': description,
            'remarks_raw': remarks,
            'online_offline': 'Online' if 'UPI' in remarks.upper() else 'Offline',
            'bank_mode': 'GPay' if 'UPI' in remarks.upper() else '',
        })
    return transactions

def insert_transactions(transactions, account_id=DEFAULT_ACCOUNT_ID):
    conn = sqlite3.connect(DB_PATH)
    inserted = 0
    for tx in transactions:
        try:
            cat_id = get_category_id(conn, USER_ID, tx['category_name'], tx['type'])
            tx_id = str(uuid.uuid4()).upper()
            conn.execute("""
                INSERT INTO Transactions 
                (Id, AccountId, Amount, BankMode, CategoryId, Date, Description, IsAutoDebit, IsMonitor, OnlineOffline, Type, UserId)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (tx_id, account_id, tx['amount'], tx['bank_mode'], cat_id, tx['date'], tx['description'], 0, 0, tx['online_offline'], tx['type'], USER_ID))
            inserted += 1
            print(f"  [OK] Rs.{tx['amount']:>10,.2f}  {tx['type']:<10}  {tx['category_name']:<15}  {tx['description']}")
        except Exception as e:
            print(f"  [ERROR] {e}")
    conn.commit()
    conn.close()
    return inserted

if __name__ == '__main__':
    xlsx_path = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\TestData.xlsx'
    print(f"Parsing {xlsx_path}...")
    transactions = parse_xlsx(xlsx_path)
    print(f"Found {len(transactions)} valid transactions.")
    if transactions:
        print("Inserting...")
        count = insert_transactions(transactions)
        print(f"SUCCESS: Imported {count} transactions!")
