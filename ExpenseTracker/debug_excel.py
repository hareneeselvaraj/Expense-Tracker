"""
Test what MiniExcel sees vs openpyxl.
MiniExcel in C# might see different headers than openpyxl in Python.
Let's simulate the same logic the C# code uses.
"""
import openpyxl, codecs

XLSX_PATH = r'C:\Users\haren\Desktop\Expense_Tracker\Expense-Tracker\TestData.xlsx'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

with codecs.open('miniexcel_debug.txt', 'w', 'utf-8') as f:
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else '')
    
    f.write(f"Headers: {headers}\n\n")
    
    # Check if fuzzy matching would work
    for partial in ['Date', 'Remarks', 'Withdrawal', 'Deposit']:
        matches = [h for h in headers if partial.lower() in h.lower()]
        f.write(f"Fuzzy '{partial}': {matches}\n")
    
    f.write("\n--- First 3 data rows ---\n")
    for r in range(2, min(ws.max_row + 1, 5)):
        f.write(f"\nRow {r}:\n")
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=col).value
            f.write(f"  {headers[col-1]}: {repr(v)} (type: {type(v).__name__})\n")

print("Done - check miniexcel_debug.txt")
